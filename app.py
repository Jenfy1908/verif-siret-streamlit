import streamlit as st
import pandas as pd
import requests
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ----------------------------------------------------------
# CONFIGURATION
# ----------------------------------------------------------
API_KEY = st.secrets["API_KEY"]
HEADERS = {"X-INSEE-Api-Key-Integration": API_KEY}
API_URL = "https://api.insee.fr/api-sirene/3.11/siret/"

st.set_page_config(page_title="Vérification SIRET", page_icon="🏢")
st.title("🏢 Vérificateur SIRET - API INSEE")

# ----------------------------------------------------------
# HELPERS
# ----------------------------------------------------------
def normalize_siret(s) -> str:
    """Garde uniquement les chiffres."""
    return "".join(c for c in str(s) if c.isdigit())

def statut_from_etat(etat: str) -> str:
    if etat == "A":
        return "Actif"
    if etat == "F":
        return "Fermé"
    return f"Inconnu ({etat})"

def fill_for_statut(statut: str) -> PatternFill:
    """
    Couleurs Excel:
    - Actif  -> vert
    - Fermé  -> rouge
    - Autre  -> orange
    """
    s = (statut or "").lower()
    if "actif" in s:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    if "fermé" in s or "ferme" in s:
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# ----------------------------------------------------------
# UPLOAD CSV
# ----------------------------------------------------------
uploaded_file = st.file_uploader("📂 Importer fichier CSV (colonne 'siret')", type=["csv"])

if uploaded_file:
    df_in = pd.read_csv(uploaded_file, dtype=str)

    if "siret" not in df_in.columns:
        st.error("❌ Le fichier doit contenir une colonne 'siret'")
        st.stop()

    sirets = df_in["siret"].dropna().tolist()
    st.success(f"✅ {len(sirets)} SIRET détectés")

    if st.button("🚀 Lancer la vérification"):
        results = []
        progress = st.progress(0)
        status = st.empty()

        for i, s in enumerate(sirets, start=1):
            siret = normalize_siret(s)
            url = f"{API_URL}{siret}"

            while True:
                r = requests.get(url, headers=HEADERS)

                if r.status_code == 200:
                    etat = (
                        r.json()
                        .get("etablissement", {})
                        .get("periodesEtablissement", [{}])[0]
                        .get("etatAdministratifEtablissement", "INCONNU")
                    )
                    statut = statut_from_etat(etat)

                elif r.status_code == 404:
                    statut = "Inexistant"

                elif r.status_code == 429:
                    status.warning("⚠️ Limite API atteinte — pause 15s…")
                    time.sleep(15)
                    continue

                else:
                    statut = f"Erreur ({r.status_code})"

                results.append({"SIRET": siret, "Statut": statut})
                progress.progress(i / len(sirets))
                status.text(f"{i}/{len(sirets)} : {siret} → {statut}")
                break

            time.sleep(0.3)

        df_res = pd.DataFrame(results)
        st.success("✅ Vérification terminée")

        # ----------------------------------------------------------
        # EXPORT EXCEL STYLÉ (colorer SIRET + Statut)
        # ----------------------------------------------------------
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_res.to_excel(writer, index=False, sheet_name="Résultats")

        output.seek(0)
        wb = load_workbook(output)
        ws = wb["Résultats"]

        # Colorer colonnes A (SIRET) ET B (Statut) selon le statut (colonne B)
        for row in range(2, ws.max_row + 1):  # 1 = header
            statut_value = ws[f"B{row}"].value
            fill = fill_for_statut(str(statut_value) if statut_value is not None else "")

            ws[f"A{row}"].fill = fill  # <-- SIRET coloré aussi
            ws[f"B{row}"].fill = fill  # <-- Statut coloré

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.download_button(
            "📥 Télécharger les résultats (Excel)",
            final_output,
            file_name="resultats_siret.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

else:
    st.info("🕮 Chargez un fichier CSV pour commencer.")
